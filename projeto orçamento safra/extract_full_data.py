#!/usr/bin/env python3
"""
Extração Completa de Dados do Excel

Exporta TODOS os dados de cada aba para:
- CSV por aba
- JSON estruturado completo
- Excel limpo (sem merges)

Diferente do ooxml_profile.py que faz amostragem, este extrai 100% dos dados.
"""

import openpyxl
import json
import csv
import argparse
import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional
from collections import defaultdict


class FullDataExtractor:
    """Extrai dados completos do Excel"""
    
    def __init__(self, excel_path: Path, output_dir: Path):
        self.excel_path = excel_path
        self.output_dir = output_dir
        self.logger = logging.getLogger(__name__)
        
        # Criar diretório de saída
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Subdiretórios
        self.csv_dir = self.output_dir / "csv"
        self.csv_dir.mkdir(exist_ok=True)
    
    def extract_all(self):
        """Extrai tudo"""
        self.logger.info(f"Carregando workbook: {self.excel_path}")
        
        # Carregar com data_only=True para obter valores calculados
        wb_values = openpyxl.load_workbook(self.excel_path, data_only=True)
        # Carregar com data_only=False para obter fórmulas
        wb_formulas = openpyxl.load_workbook(self.excel_path, data_only=False)
        
        workbook_data = {
            'metadata': {
                'source_file': str(self.excel_path),
                'extracted_at': datetime.now().isoformat(),
                'total_sheets': len(wb_values.sheetnames)
            },
            'sheets': {}
        }
        
        for sheet_name in wb_values.sheetnames:
            self.logger.info(f"Extraindo aba: {sheet_name}")
            
            sheet_values = wb_values[sheet_name]
            sheet_formulas = wb_formulas[sheet_name]
            
            sheet_data = self.extract_sheet(sheet_values, sheet_formulas, sheet_name)
            workbook_data['sheets'][sheet_name] = sheet_data
            
            # Exportar CSV desta aba
            self.export_sheet_to_csv(sheet_data, sheet_name)
        
        # Salvar JSON completo
        json_path = self.output_dir / "complete_data.json"
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(workbook_data, f, indent=2, ensure_ascii=False, default=str)
        
        self.logger.info(f"JSON completo salvo em: {json_path}")
        
        return workbook_data
    
    def extract_sheet(self, sheet_values, sheet_formulas, sheet_name: str) -> Dict:
        """Extrai dados completos de uma aba"""
        
        # Dimensões
        max_row = sheet_values.max_row
        max_col = sheet_values.max_column
        
        self.logger.info(f"  Dimensões: {max_row} linhas × {max_col} colunas")
        
        # Extrair todas as células
        rows_data = []
        formulas_found = []
        merged_cells = []
        
        # Células mescladas
        for merge_range in sheet_values.merged_cells.ranges:
            merged_cells.append(str(merge_range))
        
        # Iterar por todas as linhas
        for row_idx, (row_values, row_formulas) in enumerate(zip(
            sheet_values.iter_rows(min_row=1, max_row=max_row),
            sheet_formulas.iter_rows(min_row=1, max_row=max_row)
        ), start=1):
            
            row_data = []
            
            for col_idx, (cell_val, cell_form) in enumerate(zip(row_values, row_formulas), start=1):
                
                # Valor da célula
                value = cell_val.value
                
                # Verificar se tem fórmula
                formula = None
                if cell_form.value and isinstance(cell_form.value, str) and cell_form.value.startswith('='):
                    formula = cell_form.value
                    formulas_found.append({
                        'cell': cell_form.coordinate,
                        'formula': formula,
                        'value': value
                    })
                
                # Formato numérico
                number_format = cell_val.number_format if hasattr(cell_val, 'number_format') else None
                
                # Dados da célula
                cell_data = {
                    'value': value,
                    'coordinate': cell_val.coordinate
                }
                
                if formula:
                    cell_data['formula'] = formula
                
                if number_format and number_format != 'General':
                    cell_data['format'] = number_format
                
                # Detectar tipo
                if value is not None:
                    cell_data['type'] = type(value).__name__
                
                row_data.append(cell_data)
            
            rows_data.append(row_data)
        
        # Estatísticas
        total_cells = max_row * max_col
        filled_cells = sum(1 for row in rows_data for cell in row if cell['value'] is not None)
        
        return {
            'name': sheet_name,
            'dimensions': {
                'rows': max_row,
                'columns': max_col,
                'total_cells': total_cells,
                'filled_cells': filled_cells,
                'fill_rate': (filled_cells / total_cells * 100) if total_cells > 0 else 0
            },
            'data': rows_data,
            'formulas': formulas_found,
            'merged_cells': merged_cells,
            'formula_count': len(formulas_found),
            'merge_count': len(merged_cells)
        }
    
    def export_sheet_to_csv(self, sheet_data: Dict, sheet_name: str):
        """Exporta aba para CSV"""
        
        # Sanitizar nome do arquivo
        safe_name = "".join(c for c in sheet_name if c.isalnum() or c in (' ', '-', '_')).strip()
        csv_path = self.csv_dir / f"{safe_name}.csv"
        
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Escrever dados
            for row in sheet_data['data']:
                # Extrair apenas valores
                row_values = [cell['value'] if cell['value'] is not None else '' for cell in row]
                writer.writerow(row_values)
        
        self.logger.info(f"  CSV salvo: {csv_path}")
    
    def generate_summary_report(self, workbook_data: Dict):
        """Gera relatório resumido em Markdown"""
        
        md_lines = [
            "# Relatório de Extração Completa de Dados",
            "",
            f"**Arquivo:** {workbook_data['metadata']['source_file']}",
            f"**Extraído em:** {workbook_data['metadata']['extracted_at']}",
            "",
            "## Resumo por Aba",
            "",
            "| Aba | Linhas | Colunas | Células Total | Preenchidas | Taxa | Fórmulas | Merges |",
            "|-----|--------|---------|---------------|-------------|------|----------|--------|"
        ]
        
        for sheet_name, sheet_data in workbook_data['sheets'].items():
            dim = sheet_data['dimensions']
            md_lines.append(
                f"| {sheet_name} | {dim['rows']} | {dim['columns']} | "
                f"{dim['total_cells']} | {dim['filled_cells']} | "
                f"{dim['fill_rate']:.1f}% | {sheet_data['formula_count']} | "
                f"{sheet_data['merge_count']} |"
            )
        
        md_lines.extend([
            "",
            "## Arquivos Gerados",
            "",
            "### CSVs por Aba",
            f"- Diretório: `{self.csv_dir}`",
            ""
        ])
        
        for sheet_name in workbook_data['sheets'].keys():
            safe_name = "".join(c for c in sheet_name if c.isalnum() or c in (' ', '-', '_')).strip()
            md_lines.append(f"- `{safe_name}.csv`")
        
        md_lines.extend([
            "",
            "### JSON Completo",
            "- `complete_data.json` - Todos os dados estruturados",
            "",
            "## Próximos Passos",
            "",
            "1. Revisar os CSVs gerados",
            "2. Validar integridade dos dados",
            "3. Importar para banco de dados",
            "4. Desenvolver API e frontend",
            ""
        ])
        
        # Salvar relatório
        report_path = self.output_dir / "EXTRACTION_REPORT.md"
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(md_lines))
        
        self.logger.info(f"Relatório salvo: {report_path}")


def main():
    """Função principal"""
    parser = argparse.ArgumentParser(
        description="Extração completa de dados do Excel"
    )
    
    parser.add_argument(
        'excel_file',
        type=str,
        help='Arquivo Excel (.xlsx)'
    )
    
    parser.add_argument(
        '-o', '--output',
        type=str,
        default='./extracted_data',
        help='Diretório de saída (default: ./extracted_data)'
    )
    
    parser.add_argument(
        '--log-level',
        type=str,
        default='INFO',
        choices=['DEBUG', 'INFO', 'WARNING', 'ERROR'],
        help='Nível de log'
    )
    
    args = parser.parse_args()
    
    # Setup logging
    logging.basicConfig(
        level=getattr(logging, args.log_level),
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    
    logger = logging.getLogger(__name__)
    
    # Validar entrada
    excel_path = Path(args.excel_file)
    if not excel_path.exists():
        logger.error(f"Arquivo não encontrado: {excel_path}")
        return 1
    
    output_dir = Path(args.output)
    
    # Extrair
    try:
        extractor = FullDataExtractor(excel_path, output_dir)
        workbook_data = extractor.extract_all()
        extractor.generate_summary_report(workbook_data)
        
        print(f"\n✅ Extração concluída!")
        print(f"📁 Dados salvos em: {output_dir.absolute()}")
        print(f"📊 CSVs em: {extractor.csv_dir.absolute()}")
        print(f"📄 JSON completo: {output_dir / 'complete_data.json'}")
        print(f"📋 Relatório: {output_dir / 'EXTRACTION_REPORT.md'}")
        
        return 0
        
    except Exception as e:
        logger.error(f"Erro durante extração: {e}")
        logger.exception("Detalhes:")
        return 1


if __name__ == "__main__":
    exit(main())
