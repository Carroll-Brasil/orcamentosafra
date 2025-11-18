#!/usr/bin/env python3
"""
Script para Desmesclar Células e Propagar Valores

Resolve o problema de células mescladas do Excel:
- Lê arquivo Excel (.xlsx)
- Identifica células mescladas
- Desmescla e replica o valor para todas as células do range
- Salva nova versão "limpa" do arquivo
"""

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import argparse
import logging
from pathlib import Path
from copy import copy


def setup_logging(level='INFO'):
    """Configura logging"""
    logging.basicConfig(
        level=getattr(logging, level),
        format='%(asctime)s - %(levelname)s - %(message)s'
    )


def unmerge_and_fill(sheet):
    """
    Desmescla células e propaga valores
    
    Args:
        sheet: Worksheet do openpyxl
    
    Returns:
        int: Número de merges processados
    """
    logger = logging.getLogger(__name__)
    
    # Obter lista de células mescladas
    merged_cells = list(sheet.merged_cells.ranges)
    merge_count = len(merged_cells)
    
    logger.info(f"Encontradas {merge_count} células mescladas na aba '{sheet.title}'")
    
    for merge_range in merged_cells:
        # Obter célula inicial (que contém o valor)
        min_col, min_row, max_col, max_row = merge_range.bounds
        
        # Obter valor e estilo da primeira célula
        source_cell = sheet.cell(min_row, min_col)
        value = source_cell.value
        
        # Copiar estilo
        style_copy = None
        if source_cell.has_style:
            style_copy = copy(source_cell)
        
        logger.debug(f"Processando merge {merge_range.coord}: '{value}'")
        
        # Desmesclar
        sheet.unmerge_cells(str(merge_range))
        
        # Propagar valor e estilo para todas as células do range
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = sheet.cell(row, col)
                cell.value = value
                
                # Copiar estilo se existir
                if style_copy and cell != source_cell:
                    if style_copy.font:
                        cell.font = copy(style_copy.font)
                    if style_copy.border:
                        cell.border = copy(style_copy.border)
                    if style_copy.fill:
                        cell.fill = copy(style_copy.fill)
                    if style_copy.number_format:
                        cell.number_format = copy(style_copy.number_format)
                    if style_copy.protection:
                        cell.protection = copy(style_copy.protection)
                    if style_copy.alignment:
                        cell.alignment = copy(style_copy.alignment)
    
    return merge_count


def process_workbook(input_file: Path, output_file: Path):
    """
    Processa workbook completo
    
    Args:
        input_file: Caminho do arquivo Excel original
        output_file: Caminho do arquivo de saída
    """
    logger = logging.getLogger(__name__)
    
    logger.info(f"Carregando workbook: {input_file}")
    wb = openpyxl.load_workbook(input_file)
    
    total_merges = 0
    
    for sheet_name in wb.sheetnames:
        logger.info(f"Processando aba: {sheet_name}")
        sheet = wb[sheet_name]
        
        merge_count = unmerge_and_fill(sheet)
        total_merges += merge_count
    
    # Salvar
    logger.info(f"Salvando workbook processado: {output_file}")
    wb.save(output_file)
    
    logger.info(f"✅ Concluído! Total de {total_merges} células mescladas processadas")
    
    return total_merges


def main():
    """Função principal"""
    parser = argparse.ArgumentParser(
        description="Desmescla células do Excel e propaga valores"
    )
    
    parser.add_argument(
        'input_file',
        type=str,
        help='Arquivo Excel de entrada (.xlsx)'
    )
    
    parser.add_argument(
        '-o', '--output',
        type=str,
        default=None,
        help='Arquivo de saída (default: {input}_unmerged.xlsx)'
    )
    
    parser.add_argument(
        '--log-level',
        type=str,
        default='INFO',
        choices=['DEBUG', 'INFO', 'WARNING', 'ERROR'],
        help='Nível de log'
    )
    
    args = parser.parse_args()
    
    # Setup
    setup_logging(args.log_level)
    logger = logging.getLogger(__name__)
    
    # Validar entrada
    input_path = Path(args.input_file)
    if not input_path.exists():
        logger.error(f"Arquivo não encontrado: {input_path}")
        return 1
    
    # Definir saída
    if args.output:
        output_path = Path(args.output)
    else:
        output_path = input_path.parent / f"{input_path.stem}_unmerged{input_path.suffix}"
    
    # Processar
    try:
        process_workbook(input_path, output_path)
        print(f"\n✅ Arquivo processado salvo em: {output_path.absolute()}")
        return 0
    except Exception as e:
        logger.error(f"Erro durante processamento: {e}")
        logger.exception("Detalhes:")
        return 1


if __name__ == "__main__":
    exit(main())
