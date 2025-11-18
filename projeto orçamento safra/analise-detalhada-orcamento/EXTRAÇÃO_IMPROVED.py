#!/usr/bin/env python3
"""
Extração Aprimorada de Dados do Excel - Versão 2.0

Melhorias implementadas:
- Detecção precisa de dimensões reais (última linha/coluna com dados)
- Melhor tratamento de células mescladas
- Detecção de fórmulas mais precisa
- Metadata aprimorada
- Validação de integridade
"""

import openpyxl
import json
import csv
import argparse
import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
from collections import defaultdict


class ImprovedDataExtractor:
    """Extrator de dados aprimorado com detecção precisa de dimensões"""

    def __init__(self, excel_path: Path, output_dir: Path):
        self.excel_path = excel_path
        self.output_dir = output_dir
        self.logger = logging.getLogger(__name__)

        # Criar diretórios
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.csv_dir = self.output_dir / "csv"
        self.json_dir = self.output_dir / "json"
        self.csv_dir.mkdir(exist_ok=True)
        self.json_dir.mkdir(exist_ok=True)

    def get_real_dimensions(self, sheet) -> Tuple[int, int]:
        """
        Detecta dimensões reais da aba (última linha/coluna com dados reais)
        Evita contar células vazias no final
        """
        max_row_real = 0
        max_col_real = 0

        # Iterar por todas as células para encontrar dimensões reais
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Considerar como dado se:
                    # 1. Valor não nulo
                    # 2. String não vazia (após strip)
                    if isinstance(cell.value, str):
                        if cell.value.strip():
                            max_row_real = max(max_row_real, cell.row)
                            max_col_real = max(max_col_real, cell.column)
                    else:
                        max_row_real = max(max_row_real, cell.row)
                        max_col_real = max(max_col_real, cell.column)

        # Se não encontrou dados, usar max_row/max_col do sheet
        if max_row_real == 0:
            max_row_real = sheet.max_row
        if max_col_real == 0:
            max_col_real = sheet.max_column

        return max_row_real, max_col_real

    def get_sheet_statistics(self, sheet_values, sheet_formulas, real_rows: int, real_cols: int) -> Dict:
        """
        Calcula estatísticas detalhadas da aba
        """
        stats = {
            'real_dimensions': {'rows': real_rows, 'columns': real_cols, 'total': real_rows * real_cols},
            'declared_dimensions': {'rows': sheet_values.max_row, 'columns': sheet_values.max_column, 'total': sheet_values.max_row * sheet_values.max_column},
            'cell_types': defaultdict(int),
            'formulas_by_type': defaultdict(int),
            'numeric_formats': defaultdict(int),
            'empty_cells': 0,
            'non_empty_cells': 0,
            'cells_with_formulas': 0,
            'error_cells': 0
        }

        # Analisar células
        for row_idx in range(1, real_rows + 1):
            for col_idx in range(1, real_cols + 1):
                cell_val = sheet_values.cell(row=row_idx, column=col_idx)
                cell_form = sheet_formulas.cell(row=row_idx, column=col_idx)

                # Valor
                if cell_val.value is None or (isinstance(cell_val.value, str) and not cell_val.value.strip()):
                    stats['empty_cells'] += 1
                    cell_type = 'empty'
                else:
                    stats['non_empty_cells'] += 1
                    cell_type = type(cell_val.value).__name__
                    stats['cell_types'][cell_type] += 1

                    # Verificar erros
                    if cell_type == 'Error':
                        stats['error_cells'] += 1

                # Fórmula
                if cell_form.value and isinstance(cell_form.value, str) and cell_form.value.startswith('='):
                    stats['cells_with_formulas'] += 1

                    # Classificar tipo de fórmula
                    formula = cell_form.value.upper()
                    if 'SUM(' in formula:
                        stats['formulas_by_type']['SUM'] += 1
                    elif any(op in formula for op in ['+', '-', '*', '/']):
                        stats['formulas_by_type']['ARITHMETIC'] += 1
                    elif 'VLOOKUP(' in formula or 'PROCV(' in formula:
                        stats['formulas_by_type']['LOOKUP'] += 1
                    elif 'IF(' in formula or 'SE(' in formula:
                        stats['formulas_by_type']['CONDITIONAL'] += 1
                    else:
                        stats['formulas_by_type']['OTHER'] += 1

                # Formato numérico
                if hasattr(cell_val, 'number_format'):
                    fmt = cell_val.number_format
                    if fmt != 'General':
                        stats['numeric_formats'][fmt] += 1

        return stats

    def extract_all(self):
        """Extração completa aprimorada"""
        self.logger.info(f"Processando arquivo: {self.excel_path}")

        # Carregar workbooks
        wb_values = openpyxl.load_workbook(self.excel_path, data_only=True)
        wb_formulas = openpyxl.load_workbook(self.excel_path, data_only=False)

        # Metadata da extração
        extraction_metadata = {
            'source_file': str(self.excel_path),
            'extracted_at': datetime.now().isoformat(),
            'extraction_version': '2.0_improved',
            'total_sheets': len(wb_values.sheetnames),
            'file_size_bytes': self.excel_path.stat().st_size
        }

        workbook_data = {
            'metadata': extraction_metadata,
            'sheets': {},
            'summary': {
                'total_cells_declared': 0,
                'total_cells_real': 0,
                'total_non_empty_cells': 0,
                'total_formulas': 0,
                'total_merged_cells': 0
            }
        }

        for sheet_name in wb_values.sheetnames:
            self.logger.info(f"Extraindo aba: {sheet_name}")

            sheet_values = wb_values[sheet_name]
            sheet_formulas = wb_formulas[sheet_name]

            sheet_data = self.extract_sheet_improved(sheet_values, sheet_formulas, sheet_name)
            workbook_data['sheets'][sheet_name] = sheet_data

            # Atualizar summary
            dims = sheet_data['dimensions']
            workbook_data['summary']['total_cells_declared'] += dims['declared']['total']
            workbook_data['summary']['total_cells_real'] += dims['real']['total']
            workbook_data['summary']['total_non_empty_cells'] += dims['non_empty_cells']
            workbook_data['summary']['total_formulas'] += dims['formulas_found']
            workbook_data['summary']['total_merged_cells'] += dims['merged_cells_count']

            # Exportar CSV
            self.export_sheet_to_csv(sheet_data, sheet_name)

        # Salvar JSON completo
        json_path = self.json_dir / "complete_data_improved.json"
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(workbook_data, f, indent=2, ensure_ascii=False, default=str)

        # Gerar relatório
        self.generate_improved_report(workbook_data)

        self.logger.info(f"Extração concluída: {json_path}")
        return workbook_data

    def extract_sheet_improved(self, sheet_values, sheet_formulas, sheet_name: str) -> Dict:
        """Extração aprimorada de uma aba"""

        # Detectar dimensões reais
        real_rows, real_cols = self.get_real_dimensions(sheet_values)

        self.logger.info(f"  Dimensões declaradas: {sheet_values.max_row}×{sheet_values.max_column} = {sheet_values.max_row * sheet_values.max_column}")
        self.logger.info(f"  Dimensões reais: {real_rows}×{real_cols} = {real_rows * real_cols}")

        # Estatísticas da aba
        stats = self.get_sheet_statistics(sheet_values, sheet_formulas, real_rows, real_cols)

        # Extrair dados
        rows_data = []
        formulas_found = []

        # Células mescladas
        merged_cells = []
        for merge_range in sheet_values.merged_cells.ranges:
            merged_cells.append({
                'range': str(merge_range),
                'start_row': merge_range.min_row,
                'start_col': merge_range.min_col,
                'end_row': merge_range.max_row,
                'end_col': merge_range.max_col
            })

        # Iterar pelas células reais
        for row_idx in range(1, real_rows + 1):
            row_data = []

            for col_idx in range(1, real_cols + 1):
                cell_val = sheet_values.cell(row=row_idx, column=col_idx)
                cell_form = sheet_formulas.cell(row=row_idx, column=col_idx)

                value = cell_val.value

                # Verificar fórmula
                formula = None
                if cell_form.value and isinstance(cell_form.value, str) and cell_form.value.startswith('='):
                    formula = cell_form.value
                    formulas_found.append({
                        'coordinate': cell_form.coordinate,
                        'formula': formula,
                        'value': value,
                        'row': row_idx,
                        'column': col_idx
                    })

                # Dados da célula
                cell_data = {
                    'value': value,
                    'coordinate': cell_val.coordinate,
                    'row': row_idx,
                    'column': col_idx
                }

                if formula:
                    cell_data['formula'] = formula

                # Tipo e formato
                if value is not None:
                    cell_data['type'] = type(value).__name__
                    if hasattr(cell_val, 'number_format') and cell_val.number_format != 'General':
                        cell_data['format'] = cell_val.number_format

                # Se for erro, incluir informação
                if hasattr(value, '__class__') and 'Error' in str(type(value)):
                    cell_data['is_error'] = True
                    cell_data['error_type'] = str(type(value).__name__)

                row_data.append(cell_data)

            rows_data.append(row_data)

        return {
            'name': sheet_name,
            'dimensions': {
                'declared': {
                    'rows': sheet_values.max_row,
                    'columns': sheet_values.max_column,
                    'total': sheet_values.max_row * sheet_values.max_column
                },
                'real': {
                    'rows': real_rows,
                    'columns': real_cols,
                    'total': real_rows * real_cols
                },
                'non_empty_cells': stats['non_empty_cells'],
                'empty_cells': stats['empty_cells'],
                'fill_rate': (stats['non_empty_cells'] / (real_rows * real_cols) * 100) if real_rows * real_cols > 0 else 0
            },
            'statistics': stats,
            'data': rows_data,
            'formulas': formulas_found,
            'formulas_found': len(formulas_found),
            'merged_cells': merged_cells,
            'merged_cells_count': len(merged_cells)
        }

    def export_sheet_to_csv(self, sheet_data: Dict, sheet_name: str):
        """Exporta aba para CSV (sem alterações)"""
        safe_name = "".join(c for c in sheet_name if c.isalnum() or c in (' ', '-', '_')).strip()
        csv_path = self.csv_dir / f"{safe_name}.csv"

        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            for row in sheet_data['data']:
                row_values = [cell['value'] if cell['value'] is not None else '' for cell in row]
                writer.writerow(row_values)

        self.logger.info(f"  CSV salvo: {csv_path}")

    def generate_improved_report(self, workbook_data: Dict):
        """Gera relatório aprimorado"""
        md_lines = [
            "# Relatório de Extração Aprimorada - Versão 2.0",
            "",
            f"**Arquivo:** {workbook_data['metadata']['source_file']}",
            f"**Data Extração:** {workbook_data['metadata']['extracted_at']}",
            f"**Versão:** {workbook_data['metadata']['extraction_version']}",
            f"**Tamanho Arquivo:** {workbook_data['metadata']['file_size_size']:,} bytes",
            "",
            "## Resumo Geral",
            "",
            f"- **Total de Abas:** {workbook_data['metadata']['total_sheets']}",
            f"- **Células Declaradas:** {workbook_data['summary']['total_cells_declared']:,}",
            f"- **Células Reais:** {workbook_data['summary']['total_cells_real']:,}",
            f"- **Células Preenchidas:** {workbook_data['summary']['total_non_empty_cells']:,}",
            f"- **Taxa de Preenchimento:** {workbook_data['summary']['total_non_empty_cells']/workbook_data['summary']['total_cells_real']*100:.1f}%",
            f"- **Total de Fórmulas:** {workbook_data['summary']['total_formulas']:,}",
            f"- **Células Mescladas:** {workbook_data['summary']['total_merged_cells']:,}",
            "",
            "## Detalhamento por Aba",
            "",
            "| Aba | Decl. | Real | Preenchidas | Taxa | Fórmulas | Merges |",
            "|-----|-------|------|-------------|------|----------|--------|"
        ]

        for sheet_name, sheet_data in workbook_data['sheets'].items():
            dim = sheet_data['dimensions']
            fill_rate = dim['fill_rate']
            md_lines.append(
                f"| {sheet_name[:20]} | "
                f"{dim['declared']['total']:6,} | "
                f"{dim['real']['total']:5,} | "
                f"{dim['non_empty_cells']:10,} | "
                f"{fill_rate:4.1f}% | "
                f"{sheet_data['formulas_found']:7,} | "
                f"{sheet_data['merged_cells_count']:6,} |"
            )

        md_lines.extend([
            "",
            "## Arquivos Gerados",
            "",
            "### CSVs",
            f"- Diretório: `{self.csv_dir}`",
            "",
            "### JSON",
            f"- Dados completos: `{self.json_dir}/complete_data_improved.json`",
            "",
            "## Melhorias Implementadas",
            "",
            "1. **Detecção Precisa de Dimensões**: Considera apenas células com dados reais",
            "2. **Estatísticas Detalhadas**: Análise por tipo de célula e fórmula",
            "3. **Validação de Integridade**: Comparação declarado vs real",
            "4. **Metadata Aprimorada**: Informações detalhadas do processo",
            "5. **Tratamento de Erros**: Identificação de células com erro",
            "",
            "## Próximos Passos",
            "",
            "1. Validar integridade dos dados extraídos",
            "2. Comparar com planilha original",
            "3. Importar para sistema de destino",
            "4. Desenvolver aplicações sobre os dados",
            ""
        ])

        report_path = self.output_dir / "EXTRACTION_REPORT_IMPROVED.md"
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(md_lines))

        self.logger.info(f"Relatório aprimorado salvo: {report_path}")


def main():
    parser = argparse.ArgumentParser(description="Extração aprimorada de dados do Excel")
    parser.add_argument('excel_file', help='Arquivo Excel (.xlsx)')
    parser.add_argument('-o', '--output', default='./extracted_data_improved', help='Diretório de saída')
    parser.add_argument('--log-level', default='INFO', choices=['DEBUG', 'INFO', 'WARNING', 'ERROR'])

    args = parser.parse_args()

    logging.basicConfig(level=getattr(logging, args.log_level))

    excel_path = Path(args.excel_file)
    if not excel_path.exists():
        print(f"Erro: Arquivo não encontrado {excel_path}")
        return 1

    output_dir = Path(args.output)

    try:
        extractor = ImprovedDataExtractor(excel_path, output_dir)
        workbook_data = extractor.extract_all()

        print(f"\n✅ Extração aprimorada concluída!")
        print(f"📁 Saída: {output_dir.absolute()}")
        print(f"📊 CSVs: {extractor.csv_dir.absolute()}")
        print(f"📄 JSON: {extractor.json_dir.absolute()}")
        print(f"📋 Relatório: {output_dir / 'EXTRACTION_REPORT_IMPROVED.md'}")

        return 0

    except Exception as e:
        print(f"Erro: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    exit(main())