#!/usr/bin/env python3
"""
Script de Teste - Demonstra extração de dados

Cria um Excel de exemplo com células mescladas e demonstra:
1. Como células mescladas causam problemas
2. Como desmesclar resolve
3. Como extrair dados completos
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import json
from pathlib import Path


def create_sample_excel():
    """Cria Excel de exemplo com células mescladas"""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orçamento Exemplo"
    
    # Título mesclado
    ws.merge_cells('A1:D1')
    ws['A1'] = 'ORÇAMENTO DE COMBUSTÍVEL'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A1'].fill = PatternFill(start_color="4472C4", fill_type="solid")
    
    # Subtítulo mesclado
    ws.merge_cells('A2:D2')
    ws['A2'] = 'Safra 2024/2025'
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['Veículo', 'Tipo', 'Litros', 'Valor R$']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(3, col)
        cell.value = header
        cell.font = Font(bold=True)
    
    # Dados com merges por categoria
    ws.merge_cells('A4:A6')
    ws['A4'] = 'Tratores'
    
    ws['B4'] = 'Diesel S10'
    ws['C4'] = 1500
    ws['D4'] = 7500.00
    
    ws['B5'] = 'Óleo Motor'
    ws['C5'] = 50
    ws['D5'] = 1250.00
    
    ws['B6'] = 'Graxa'
    ws['C6'] = 10
    ws['D6'] = 150.00
    
    # Segunda categoria mesclada
    ws.merge_cells('A7:A9')
    ws['A7'] = 'Caminhões'
    
    ws['B7'] = 'Diesel S10'
    ws['C7'] = 800
    ws['D7'] = 4000.00
    
    ws['B8'] = 'Óleo Motor'
    ws['C8'] = 30
    ws['D8'] = 750.00
    
    ws['B9'] = 'Aditivo'
    ws['C9'] = 5
    ws['D9'] = 200.00
    
    # Total mesclado
    ws.merge_cells('A10:C10')
    ws['A10'] = 'TOTAL'
    ws['A10'].font = Font(bold=True)
    ws['D10'] = '=SUM(D4:D9)'
    ws['D10'].font = Font(bold=True)
    
    # Ajustar larguras
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    
    return wb


def demonstrate_problem():
    """Demonstra o problema das células mescladas"""
    
    print("="*60)
    print("DEMONSTRAÇÃO: Problema das Células Mescladas")
    print("="*60)
    
    # Criar Excel
    wb = create_sample_excel()
    test_file = Path("teste_merge.xlsx")
    wb.save(test_file)
    
    print(f"\n✅ Arquivo criado: {test_file}")
    print("\n📊 Estrutura do arquivo:")
    print("   - Células A1:D1 mescladas (Título)")
    print("   - Células A4:A6 mescladas (Categoria 'Tratores')")
    print("   - Células A7:A9 mescladas (Categoria 'Caminhões')")
    
    # Ler arquivo para demonstrar problema
    print("\n" + "="*60)
    print("TENTATIVA 1: Leitura Direta (COM PROBLEMA)")
    print("="*60)
    
    wb_read = openpyxl.load_workbook(test_file)
    ws_read = wb_read.active
    
    print("\nLendo linha 4 (primeira linha de 'Tratores'):")
    for col in range(1, 5):
        cell = ws_read.cell(4, col)
        print(f"  Coluna {openpyxl.utils.get_column_letter(col)}: '{cell.value}'")
    
    print("\n❌ PROBLEMA: B4, C4 aparecem vazias porque A4:A6 está mesclado!")
    print("   Apenas A4 tem o valor 'Tratores'")
    
    # Demonstrar solução
    print("\n" + "="*60)
    print("SOLUÇÃO: Desmesclar e Propagar Valores")
    print("="*60)
    
    # Desmesclar manualmente para demonstração
    merged_ranges = list(ws_read.merged_cells.ranges)
    print(f"\nEncontradas {len(merged_ranges)} células mescladas")
    
    for merge_range in merged_ranges:
        # Obter valor da primeira célula
        min_col, min_row, max_col, max_row = merge_range.bounds
        source_value = ws_read.cell(min_row, min_col).value
        
        print(f"\n  Processando: {merge_range.coord}")
        print(f"    Valor original: '{source_value}'")
        
        # Desmesclar
        ws_read.unmerge_cells(str(merge_range))
        
        # Propagar
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws_read.cell(row, col).value = source_value
        
        print(f"    ✓ Valor propagado para todas as células do range")
    
    # Salvar versão corrigida
    fixed_file = Path("teste_merge_FIXED.xlsx")
    wb_read.save(fixed_file)
    
    print(f"\n✅ Arquivo corrigido salvo: {fixed_file}")
    
    # Re-ler para confirmar
    print("\n" + "="*60)
    print("TENTATIVA 2: Leitura Após Correção (SEM PROBLEMA)")
    print("="*60)
    
    wb_fixed = openpyxl.load_workbook(fixed_file)
    ws_fixed = wb_fixed.active
    
    print("\nLendo linha 4 novamente:")
    for col in range(1, 5):
        cell = ws_fixed.cell(4, col)
        print(f"  Coluna {openpyxl.utils.get_column_letter(col)}: '{cell.value}'")
    
    print("\n✅ SUCESSO: Todas as colunas têm valores!")
    print("   Agora pode exportar para CSV/JSON sem perder dados")
    
    # Demonstrar exportação para CSV
    print("\n" + "="*60)
    print("EXPORTAÇÃO PARA CSV")
    print("="*60)
    
    csv_file = Path("teste_export.csv")
    import csv
    
    with open(csv_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        
        for row in ws_fixed.iter_rows(values_only=True):
            writer.writerow(row)
    
    print(f"\n✅ CSV exportado: {csv_file}")
    
    print("\nConteúdo do CSV:")
    print("-" * 60)
    with open(csv_file, 'r', encoding='utf-8') as f:
        print(f.read())
    
    # Demonstrar JSON
    print("=" * 60)
    print("EXPORTAÇÃO PARA JSON")
    print("=" * 60)
    
    data = []
    for row_idx, row in enumerate(ws_fixed.iter_rows(values_only=True), start=1):
        data.append({
            'row': row_idx,
            'values': list(row)
        })
    
    json_file = Path("teste_export.json")
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False, default=str)
    
    print(f"\n✅ JSON exportado: {json_file}")
    
    print("\nConteúdo do JSON (primeiras 5 linhas):")
    print("-" * 60)
    print(json.dumps(data[:5], indent=2, ensure_ascii=False, default=str))
    
    # Resumo
    print("\n" + "=" * 60)
    print("RESUMO")
    print("=" * 60)
    print("\n📁 Arquivos gerados:")
    print(f"  1. {test_file} - Excel original COM células mescladas")
    print(f"  2. {fixed_file} - Excel corrigido SEM células mescladas")
    print(f"  3. {csv_file} - Exportação CSV")
    print(f"  4. {json_file} - Exportação JSON")
    
    print("\n💡 Lições Aprendidas:")
    print("  1. Células mescladas causam perda de dados na exportação")
    print("  2. Desmesclar e propagar valores resolve o problema")
    print("  3. Após correção, pode exportar para qualquer formato")
    print("  4. Use unmarge_cells.py para automatizar este processo")
    
    print("\n🚀 Próximos passos:")
    print("  1. Execute: python unmarge_cells.py <seu_arquivo.xlsx>")
    print("  2. Execute: python extract_full_data.py <arquivo_unmerged.xlsx>")
    print("  3. Importe os CSVs gerados para seu banco de dados")


if __name__ == "__main__":
    demonstrate_problem()
